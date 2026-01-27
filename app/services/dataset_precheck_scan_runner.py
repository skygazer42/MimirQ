"""
Dataset precheck scan runner.

This scans a local folder (before ingestion) and produces:
- per-file records (JSONL on disk) for drill-down
- an aggregated summary snapshot (stored in DB)

Security:
- local scan must be explicitly enabled (LOCAL_SCAN_ENABLED)
- scan root must be under UPLOAD_DIR or one of LOCAL_SCAN_ROOTS
- symlinks are not followed (by design)
"""

from __future__ import annotations

import contextlib
import hashlib
import io
import json
import os
import re
from collections import Counter
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional
from uuid import UUID

from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.dataset_precheck_scan import DatasetPrecheckScanRun as DBDatasetPrecheckScanRun
from app.rag.core.logging import get_logger
from app.rag.preprocessing.pii_anonymizer import anonymize_pii, find_pii_matches
from app.rag.preprocessing.secrets import redact_secrets, find_secret_matches
from app.services.dataset_profile_utils import FILE_SIZE_BINS, TEXT_LENGTH_BINS, histogram, percentile_from_sorted, safe_int

logger = get_logger("services.dataset_precheck_scan")


FINDING_KEY_REASONS: dict[str, dict[str, Any]] = {
    "parse_failed": {
        "label": "解析/读取失败",
        "severity": "error",
        "description": "文件无法读取或解析（权限/损坏/依赖缺失）。",
    },
    "pdf_scanned": {
        "label": "疑似扫描 PDF",
        "severity": "warning",
        "description": "可能需要 OCR/更强 PDF 解析链路。",
    },
    "pdf_unknown": {
        "label": "PDF 类型未知",
        "severity": "info",
        "description": "无法获取 PDF 指标（解析失败/加密/权限）。",
    },
    "pii": {
        "label": "PII 命中",
        "severity": "warning",
        "description": "命中手机号/邮箱/身份证等（抽样文本）。建议复核脱敏策略。",
    },
    "secrets": {
        "label": "密钥/Token 命中",
        "severity": "warning",
        "description": "命中疑似密钥/Token（抽样文本）。建议脱敏或隔离。",
    },
    "exact_dup": {
        "label": "完全重复候选",
        "severity": "info",
        "description": "基于 file_sha256 的完全重复候选（需开启 compute_file_hash）。",
    },
    "near_dup": {
        "label": "近重复候选（版本冲突）",
        "severity": "info",
        "description": "基于文本样本的 SimHash 近重复候选（需开启 enable_near_dup；需要人工复核）。",
    },
    "large_spreadsheet": {
        "label": "大型表格（建议结构化方案）",
        "severity": "info",
        "description": "行数过多的表格更适合 Text-to-SQL/结构化索引，而非纯向量检索。",
    },
}


TEXTLIKE_EXTS = {
    ".txt",
    ".md",
    ".rst",
    ".adoc",
    ".asciidoc",
    ".tex",
    ".yaml",
    ".yml",
    ".toml",
    ".ini",
    ".cfg",
    ".conf",
    ".env",
    ".properties",
    ".sql",
    ".log",
    ".patch",
    ".diff",
    ".csv",
    ".json",
    ".jsonl",
    ".ndjson",
    ".xml",
    ".rss",
    ".atom",
    ".graphql",
    ".gql",
    ".proto",
    ".tf",
    ".hcl",
    ".html",
    ".htm",
}


def _hash64(text: str) -> int:
    # Stable 64-bit hash for SimHash features (blake2b is fast and deterministic).
    h = hashlib.blake2b((text or "").encode("utf-8", errors="ignore"), digest_size=8).digest()
    return int.from_bytes(h, byteorder="big", signed=False)


def _simhash64(text: str, *, max_tokens: int = 2000) -> int:
    """
    Compute a simple 64-bit SimHash for near-duplicate detection (best-effort).

    We intentionally keep this dependency-free and conservative:
    - tokenize by simple regex (latin alnum + CJK)
    - cap token count to avoid large CPU spikes during scans
    """
    s = (text or "").strip()
    if not s:
        return 0

    tokens = re.findall(r"[A-Za-z0-9_\u4e00-\u9fff]{2,}", s.lower())
    if not tokens:
        return 0
    if len(tokens) > max_tokens:
        tokens = tokens[:max_tokens]

    counts = Counter(tokens)
    # 64-dim signed accumulator.
    v = [0] * 64
    for tok, w in counts.items():
        hv = _hash64(tok)
        weight = int(w)
        for i in range(64):
            if (hv >> i) & 1:
                v[i] += weight
            else:
                v[i] -= weight

    out = 0
    for i, val in enumerate(v):
        if val > 0:
            out |= 1 << i
    return int(out)


def _hamming_distance64(a: int, b: int) -> int:
    return int((int(a) ^ int(b)).bit_count())


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _parse_csv(raw: str) -> list[str]:
    parts = [p.strip() for p in str(raw or "").split(",")]
    return [p for p in parts if p]


def _assert_local_scan_enabled() -> None:
    if not bool(getattr(settings, "LOCAL_SCAN_ENABLED", False)):
        raise ValueError("local_scan_disabled")


def _assert_scan_root_allowed(root: Path) -> None:
    """
    Ensure root is within UPLOAD_DIR or one of LOCAL_SCAN_ROOTS.

    This is a safety guard against arbitrary file reads in shared deployments.
    """
    upload_root = Path(getattr(settings, "UPLOAD_DIR", "./uploads") or "./uploads").resolve(strict=False)
    allowed: list[Path] = [upload_root]
    for p in _parse_csv(str(getattr(settings, "LOCAL_SCAN_ROOTS", "") or "")):
        try:
            allowed.append(Path(p).expanduser().resolve(strict=False))
        except Exception:
            continue

    resolved = root.expanduser().resolve(strict=False)
    for base in allowed:
        try:
            resolved.relative_to(base)
            return
        except Exception:
            continue
    raise ValueError("scan_root_denied")


def _safe_hash_file(path: Path, *, algo: str = "sha256", chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.new(algo)
    with path.open("rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _read_text_sample(path: Path, *, max_bytes: int) -> tuple[str, bool]:
    """
    Best-effort read a text sample from a file.

    Returns (text, estimated), where estimated means sample-only (file too large).
    """
    size = path.stat().st_size
    max_bytes = max(1_000, int(max_bytes or 0))
    read_bytes = min(size, max_bytes)
    with path.open("rb") as f:
        buf = f.read(read_bytes)
    # Best-effort decode: utf-8 with replacement. We only need rough metrics/hits.
    text = buf.decode("utf-8", errors="ignore")
    return text, bool(size > read_bytes)


def _pdf_text_sample(path: Path, *, sample_pages: int, max_chars: int = 200_000) -> tuple[str, bool, int, list[int | None]]:
    """
    Extract a best-effort text sample from first N pages of a PDF.

    Returns (text, estimated, page_count, per_page_chars) where per_page_chars is
    a list aligned with sampled pages (None means extraction failed for that page).
    """
    try:
        import pdfplumber  # type: ignore

        with pdfplumber.open(str(path)) as pdf:
            page_count = len(pdf.pages)
            n = max(1, min(int(sample_pages or 1), page_count))
            chunks: list[str] = []
            per_page_chars: list[int | None] = []
            total = 0
            for p in pdf.pages[:n]:
                try:
                    t = p.extract_text() or ""
                except Exception:
                    t = ""
                    per_page_chars.append(None)
                    continue
                per_page_chars.append(len(t))
                if not t:
                    continue
                if total + len(t) > max_chars:
                    t = t[: max(0, max_chars - total)]
                chunks.append(t)
                total += len(t)
                if total >= max_chars:
                    break
        text = "\n".join(chunks)
        estimated = bool(page_count > n)
        return text, estimated, int(page_count), per_page_chars
    except Exception:
        return "", True, 0, []


def _build_pdf_page_breakdown(
    *,
    page_count: int,
    per_page_chars: list[int | None],
    scan_max_chars: int,
    text_min_chars: int,
) -> dict[str, Any]:
    scanned_pages = 0
    text_pages = 0
    low_density_pages = 0
    unknown_pages = 0

    sampled_pages = int(len(per_page_chars or []))
    for ch in per_page_chars or []:
        if ch is None:
            unknown_pages += 1
            continue
        chars = int(ch or 0)
        if chars <= int(scan_max_chars):
            scanned_pages += 1
        elif chars >= int(text_min_chars):
            text_pages += 1
        else:
            low_density_pages += 1

    denom = max(1, sampled_pages - unknown_pages)
    scan_ratio = float(scanned_pages) / float(denom)
    low_density_ratio = float(low_density_pages) / float(denom)

    return {
        "page_count": int(page_count or 0),
        "sampled_pages": int(sampled_pages),
        "scanned_pages": int(scanned_pages),
        "text_pages": int(text_pages),
        "low_density_pages": int(low_density_pages),
        "unknown_pages": int(unknown_pages),
        "scan_ratio": round(scan_ratio, 4),
        "low_density_ratio": round(low_density_ratio, 4),
    }


def _sanitize_display_name(rel_path: str) -> str:
    # Keep it stable and safe for HTML/JSON; do not allow control chars.
    s = str(rel_path or "").replace("\\", "/").strip()
    s = re.sub(r"[\x00-\x1f\x7f]+", "", s)
    return s[:1024] if len(s) > 1024 else s


def _xlsx_spreadsheet_stats(path: Path, *, max_sheets: int = 3) -> dict[str, Any] | None:
    """
    Best-effort spreadsheet stats for .xlsx (read-only).

    Note: This is intentionally lightweight and may be approximate for files with
    complex formatting. It should never raise.
    """
    try:
        import openpyxl  # type: ignore
    except Exception:
        return None

    wb = None
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        sheetnames = list(getattr(wb, "sheetnames", None) or [])
        sheet_count = int(len(sheetnames))
        max_rows = 0
        max_cols = 0
        merged_area = 0

        for idx, name in enumerate(sheetnames[: max(1, int(max_sheets or 1))]):
            ws = wb[name]
            r = int(getattr(ws, "max_row", 0) or 0)
            c = int(getattr(ws, "max_column", 0) or 0)
            max_rows = max(max_rows, r)
            max_cols = max(max_cols, c)

            # Only compute merged cells on the first sheet (cheap signal).
            if idx == 0:
                merged_cells = getattr(ws, "merged_cells", None)
                ranges = list(getattr(merged_cells, "ranges", None) or [])
                # Cap range count to avoid pathological files.
                for rng in ranges[:5000]:
                    try:
                        merged_area += int(rng.size)  # type: ignore[attr-defined]
                    except Exception:
                        try:
                            merged_area += int((rng.max_row - rng.min_row + 1) * (rng.max_col - rng.min_col + 1))  # type: ignore[attr-defined]
                        except Exception:
                            continue

        total_area = max(1, int(max_rows) * int(max_cols))
        merged_ratio = float(merged_area) / float(total_area)
        if merged_ratio < 0.0:
            merged_ratio = 0.0
        if merged_ratio > 1.0:
            merged_ratio = 1.0

        return {
            "row_count": int(max_rows),
            "sheet_count": int(sheet_count),
            "merged_cell_ratio": round(float(merged_ratio), 6),
            "estimated_rows": False,
        }
    except Exception:
        return None
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass


def _mask_pii_value(kind: str, raw: str) -> str:
    k = (kind or "").strip().lower()
    s = (raw or "").strip()
    if not s:
        return "[REDACTED]"
    if k == "email":
        if "@" not in s:
            return "[REDACTED]"
        local, domain = s.split("@", 1)
        head = (local[:1] + "***") if local else "***"
        return f"{head}@{domain}"
    if k == "ip":
        parts = s.split(".")
        if len(parts) == 4:
            return ".".join(parts[:3] + ["***"])
        return "[REDACTED]"
    if k in {"phone"}:
        digits = re.sub(r"[^\d]", "", s)
        if len(digits) >= 7:
            return f"{digits[:3]}****{digits[-2:]}"
        return "[REDACTED]"
    if k == "credit_card":
        digits = re.sub(r"[^\d]", "", s)
        if len(digits) >= 8:
            return f"{digits[:4]}****{digits[-4:]}"
        return "[REDACTED]"
    if k == "cn_id":
        if len(s) >= 10:
            return f"{s[:6]}********{s[-2:]}"
        return "[REDACTED]"
    if k == "ssn":
        return "***-**-****"
    return "[REDACTED]"


def _mask_secret_value(kind: str, raw: str) -> str:
    k = (kind or "").strip().lower()
    s = (raw or "").strip()
    if not s:
        return "[SECRET]"
    if k == "openai_key":
        return "sk-***"
    if k == "github_token":
        if s.startswith("ghp_"):
            return "ghp_***"
        if s.startswith("github_pat_"):
            return "github_pat_***"
        return "[SECRET]"
    if k == "aws_access_key":
        return (s[:4] + "***") if len(s) >= 4 else "[SECRET]"
    if k == "slack_token":
        # xox[baprs]-...
        prefix = s.split("-", 1)[0]
        return f"{prefix}-***" if prefix else "[SECRET]"
    if k == "bearer_token":
        return "Bearer ***"
    if k == "private_key":
        return "-----BEGIN PRIVATE KEY----- ... -----END PRIVATE KEY-----"
    return "[SECRET]"


@dataclass
class _FileRecord:
    name: str
    file_type: str
    file_size: int
    file_mtime: int = 0
    text_characters: int = 0
    estimated_text: bool = False
    pdf_scanned: Optional[bool] = None
    pdf_pages: Optional[dict[str, Any]] = None
    spreadsheet: Optional[dict[str, Any]] = None
    text_simhash64: Optional[str] = None
    pii_hits: dict[str, int] = field(default_factory=dict)
    secrets_hits: dict[str, int] = field(default_factory=dict)
    pii_samples: list[dict[str, Any]] = field(default_factory=list)
    secrets_samples: list[dict[str, Any]] = field(default_factory=list)
    file_sha256: Optional[str] = None
    findings: list[str] = field(default_factory=list)
    error_message: Optional[str] = None


def _iter_files(root: Path, *, max_files: int) -> Iterable[Path]:
    # Use os.walk to avoid following symlinks by default.
    count = 0
    for dirpath, dirnames, filenames in os.walk(str(root), topdown=True, followlinks=False):
        # Skip hidden dirs by default (can be added later via config).
        dirnames[:] = [d for d in dirnames if not d.startswith(".")]
        for fn in filenames:
            if fn.startswith("."):
                continue
            path = Path(dirpath) / fn
            if not path.is_file():
                continue
            yield path
            count += 1
            if count >= max_files:
                return


def run_dataset_precheck_scan(
    db: Session,
    *,
    tenant_id: UUID,
    dataset_id: UUID,
    scan_run_id: UUID,
) -> dict[str, Any]:
    """
    Execute a dataset precheck scan run.

    Returns a small stats dict for logs/worker result.
    """
    run = (
        db.query(DBDatasetPrecheckScanRun)
        .filter(
            DBDatasetPrecheckScanRun.id == scan_run_id,
            DBDatasetPrecheckScanRun.tenant_id == tenant_id,
            DBDatasetPrecheckScanRun.dataset_id == dataset_id,
        )
        .first()
    )
    if run is None:
        raise ValueError("scan_run_not_found")

    # Mark running.
    run.status = "running"
    run.progress = 0
    run.started_at = _now_utc()
    run.error_message = None
    db.commit()

    cfg = dict(getattr(run, "config", None) or {})
    root_path = str(cfg.get("root_path") or "").strip()
    if not root_path:
        raise ValueError("root_path_required")

    _assert_local_scan_enabled()
    root = Path(root_path)
    if not root.exists() or not root.is_dir():
        raise ValueError("root_path_not_found")
    _assert_scan_root_allowed(root)

    max_files_cap = safe_int(getattr(settings, "PRECHECK_SCAN_MAX_FILES", 20_000), default=20_000)
    requested_max_files = cfg.get("max_files")
    max_files_req = safe_int(requested_max_files, default=0)
    max_files = max_files_cap if max_files_req <= 0 else min(max_files_cap, max_files_req)

    max_total_bytes = safe_int(getattr(settings, "PRECHECK_SCAN_MAX_TOTAL_BYTES", 0), default=0)
    text_max_bytes = safe_int(
        cfg.get("text_extract_max_bytes") or getattr(settings, "PRECHECK_TEXT_EXTRACT_MAX_BYTES", 2_000_000),
        default=2_000_000,
    )
    pdf_sample_pages = safe_int(
        cfg.get("pdf_sample_pages") or getattr(settings, "PRECHECK_PDF_SAMPLE_PAGES", 3),
        default=3,
    )

    enable_pdf_quality = bool(cfg.get("enable_pdf_quality", True))
    enable_text_extract = bool(cfg.get("enable_text_extract", True))
    enable_pii = bool(cfg.get("enable_pii", False))
    enable_secrets = bool(cfg.get("enable_secrets", False))
    compute_file_hash = bool(cfg.get("compute_file_hash", False))
    redact_paths = bool(cfg.get("redact_paths", False))

    enable_pii_samples = bool(cfg.get("enable_pii_samples", False))
    pii_context_chars = safe_int(cfg.get("pii_context_chars") or 50, default=50)
    pii_context_chars = max(0, min(pii_context_chars, 500))
    pii_max_samples_per_file = safe_int(cfg.get("pii_max_samples_per_file") or 5, default=5)
    pii_max_samples_per_file = max(0, min(pii_max_samples_per_file, 50))

    enable_secrets_samples = bool(cfg.get("enable_secrets_samples", False))
    secrets_context_chars = safe_int(cfg.get("secrets_context_chars") or 50, default=50)
    secrets_context_chars = max(0, min(secrets_context_chars, 500))
    secrets_max_samples_per_file = safe_int(cfg.get("secrets_max_samples_per_file") or 5, default=5)
    secrets_max_samples_per_file = max(0, min(secrets_max_samples_per_file, 50))

    enable_near_dup = bool(cfg.get("enable_near_dup", False))
    near_dup_hamming_threshold = safe_int(cfg.get("near_dup_hamming_threshold") or 5, default=5)
    near_dup_hamming_threshold = max(0, min(near_dup_hamming_threshold, 32))
    near_dup_max_pairs = safe_int(cfg.get("near_dup_max_pairs") or 5000, default=5000)
    near_dup_max_pairs = max(0, min(near_dup_max_pairs, 200_000))

    pdf_scan_max_chars = safe_int(
        cfg.get("pdf_min_text_chars_per_page") or getattr(settings, "PRECHECK_PDF_MIN_TEXT_CHARS_PER_PAGE", 50),
        default=50,
    )
    pdf_text_min_chars = safe_int(
        cfg.get("pdf_text_chars_per_page") or getattr(settings, "PRECHECK_PDF_TEXT_CHARS_PER_PAGE", 200),
        default=200,
    )
    pdf_scan_ratio_threshold = float(
        cfg.get("pdf_scan_ratio_threshold")
        if cfg.get("pdf_scan_ratio_threshold") is not None
        else float(getattr(settings, "PRECHECK_PDF_SCAN_RATIO_THRESHOLD", 0.7) or 0.7)
    )
    spreadsheet_large_row_threshold = safe_int(
        getattr(settings, "PRECHECK_SPREADSHEET_LARGE_ROW_THRESHOLD", 5000),
        default=5000,
    )

    # Prepare artifact dir and JSONL writer.
    artifact_root = Path(getattr(settings, "UPLOAD_DIR", "./uploads") or "./uploads") / str(tenant_id) / "precheck" / str(run.id)
    artifact_root.mkdir(parents=True, exist_ok=True)
    jsonl_path = (artifact_root / "files.jsonl").resolve(strict=False)

    # Enumerate file candidates first to compute total for progress (bounded by max_files).
    allowed_exts = set(getattr(settings, "allowed_extensions_list", []) or [])
    candidates: list[Path] = []
    total_bytes_seen = 0
    for p in _iter_files(root, max_files=max_files):
        ext = p.suffix.lower()
        if ext not in allowed_exts:
            continue
        try:
            size = int(p.stat().st_size)
        except Exception:
            size = 0
        if max_total_bytes > 0 and (total_bytes_seen + size) > max_total_bytes:
            break
        candidates.append(p)
        total_bytes_seen += size

    total = len(candidates)
    if total == 0:
        run.status = "completed"
        run.progress = 100
        run.finished_at = _now_utc()
        run.summary = {
            "dataset_id": str(dataset_id),
            "scan_run_id": str(run.id),
            "generated_at": run.finished_at.isoformat(),
            "total_files": 0,
            "total_size_bytes": 0,
            "by_file_type": {},
            "file_size_histogram": [],
            "length_percentiles": {"p25": 0, "p50": 0, "p75": 0, "p90": 0, "p99": 0},
            "length_histogram": [],
            "pdf_scan": {"scanned": 0, "not_scanned": 0, "unknown": 0},
            "pdf_detection": {
                "sample_pages": int(pdf_sample_pages),
                "scan_max_chars_per_page": int(pdf_scan_max_chars),
                "text_min_chars_per_page": int(pdf_text_min_chars),
                "scan_ratio_threshold": float(pdf_scan_ratio_threshold),
            },
            "pii_hits_total": {},
            "secrets_hits_total": {},
            "findings": [
                {"key": k, "label": v.get("label", k), "severity": v.get("severity", "info"), "count": 0, "description": v.get("description")}
                for k, v in FINDING_KEY_REASONS.items()
            ],
        }
        run.artifacts = {"files_jsonl": str(jsonl_path), "root_path": "[REDACTED]" if redact_paths else str(root)}
        db.commit()
        return {"ok": True, "files": 0}

    last_progress_write = time.monotonic()

    def flush_progress(processed: int, *, force: bool = False) -> None:
        nonlocal last_progress_write
        now = time.monotonic()
        if not force and (now - last_progress_write) < 0.5:
            return
        last_progress_write = now
        pct = int((processed / max(1, total)) * 100)
        run.progress = max(0, min(100, pct))
        db.commit()

    # Aggregation accumulators.
    by_type: dict[str, int] = {}
    file_sizes: list[int] = []
    text_lengths: list[int] = []
    pii_totals: dict[str, int] = {}
    secrets_totals: dict[str, int] = {}
    pdf_scanned = 0
    pdf_not_scanned = 0
    pdf_unknown = 0
    finding_counts: dict[str, int] = {k: 0 for k in FINDING_KEY_REASONS.keys()}

    # For exact-dup finding: sha256 -> count.
    sha_counts: dict[str, int] = {}

    # For near-dup finding: (display_name, simhash64) pairs.
    simhash_entries: list[tuple[str, int]] = []

    errors = 0

    # Stream JSONL writes to avoid holding all file records in memory.
    with jsonl_path.open("w", encoding="utf-8") as jf:
        for idx, path in enumerate(candidates, start=1):
            rel = ""
            try:
                rel = str(path.relative_to(root)).replace("\\", "/")
            except Exception:
                rel = path.name

            ext = path.suffix.lower()
            file_type = ext.lstrip(".") if ext.startswith(".") else (path.suffix or "").lower().lstrip(".") or "unknown"
            try:
                st = path.stat()
                size = int(st.st_size)
                mtime = int(st.st_mtime)
            except Exception:
                size = 0
                mtime = 0

            rec = _FileRecord(
                name=_sanitize_display_name(rel) if not redact_paths else f"FILE_{idx:06d}{ext}",
                file_type=file_type or "unknown",
                file_size=int(size),
                file_mtime=int(mtime),
            )

            try:
                # Text metrics (best-effort).
                sample_text = ""
                estimated_text = False
                if enable_text_extract:
                    if ext in TEXTLIKE_EXTS:
                        sample_text, estimated_text = _read_text_sample(path, max_bytes=text_max_bytes)
                        # Estimate text length using bytes ratio (rough).
                        if sample_text:
                            if estimated_text and size > 0:
                                ratio = size / max(1, min(size, text_max_bytes))
                                rec.text_characters = int(len(sample_text) * ratio)
                                rec.estimated_text = True
                            else:
                                rec.text_characters = int(len(sample_text))
                                rec.estimated_text = False
                    elif ext == ".pdf":
                        sample_text, estimated_text, page_count, per_page_chars = _pdf_text_sample(path, sample_pages=pdf_sample_pages)
                        if sample_text:
                            if estimated_text and page_count > 0:
                                # Scale by pages (rough).
                                rec.text_characters = int(len(sample_text) * (page_count / max(1, min(page_count, pdf_sample_pages))))
                                rec.estimated_text = True
                            else:
                                rec.text_characters = int(len(sample_text))
                                rec.estimated_text = False
                        # Always attach PDF page breakdown when possible (even if text is empty).
                        if page_count > 0 and per_page_chars:
                            rec.pdf_pages = _build_pdf_page_breakdown(
                                page_count=int(page_count),
                                per_page_chars=per_page_chars,
                                scan_max_chars=int(pdf_scan_max_chars),
                                text_min_chars=int(pdf_text_min_chars),
                            )

                # Optional near-duplicate fingerprint (SimHash over extracted text sample).
                if enable_near_dup and sample_text:
                    try:
                        sim = _simhash64(sample_text)
                    except Exception:
                        sim = 0
                    if sim:
                        rec.text_simhash64 = f"{int(sim) & ((1<<64)-1):016x}"
                        simhash_entries.append((rec.name, int(sim)))

                # PDF scan detection (transparent heuristics on sampled pages).
                if enable_pdf_quality and ext == ".pdf":
                    breakdown = rec.pdf_pages if isinstance(rec.pdf_pages, dict) else None
                    scan_ratio = float(breakdown.get("scan_ratio") or 0.0) if breakdown else 0.0
                    page_count = int(breakdown.get("page_count") or 0) if breakdown else 0

                    if page_count <= 0:
                        rec.pdf_scanned = None
                        rec.findings.append("pdf_unknown")
                        finding_counts["pdf_unknown"] += 1
                        pdf_unknown += 1
                    else:
                        scanned = bool(scan_ratio >= float(pdf_scan_ratio_threshold))
                        rec.pdf_scanned = scanned
                        if scanned:
                            rec.findings.append("pdf_scanned")
                            finding_counts["pdf_scanned"] += 1
                            pdf_scanned += 1
                        else:
                            pdf_not_scanned += 1

                # Spreadsheet stats (best-effort): large tables are often better served by structured indexing / Text-to-SQL.
                if ext in {".csv", ".xlsx"}:
                    if ext == ".csv":
                        # Use the already-read text sample as a cheap proxy for row count.
                        if sample_text:
                            lines = int(sample_text.count("\n"))
                            if not sample_text.endswith("\n"):
                                lines += 1
                            row_count = max(0, lines)
                            estimated_rows = False
                            if estimated_text and size > 0:
                                ratio = size / max(1, min(size, text_max_bytes))
                                row_count = int(row_count * ratio)
                                estimated_rows = True
                            rec.spreadsheet = {
                                "row_count": int(row_count),
                                "sheet_count": 1,
                                "merged_cell_ratio": 0.0,
                                "estimated_rows": bool(estimated_rows),
                            }
                    else:
                        stats = _xlsx_spreadsheet_stats(path)
                        if isinstance(stats, dict) and stats:
                            rec.spreadsheet = stats

                    if isinstance(rec.spreadsheet, dict):
                        rows = int(rec.spreadsheet.get("row_count") or 0)
                        if spreadsheet_large_row_threshold > 0 and rows >= int(spreadsheet_large_row_threshold):
                            rec.findings.append("large_spreadsheet")
                            finding_counts["large_spreadsheet"] += 1

                # Optional PII/secrets detection on sample text.
                if sample_text and enable_pii:
                    pii = anonymize_pii(sample_text, enabled=True, mode="mask")
                    if pii.hits:
                        rec.pii_hits = {str(k): int(v) for k, v in pii.hits.items() if int(v) > 0}
                        if rec.pii_hits:
                            rec.findings.append("pii")
                            finding_counts["pii"] += 1
                            for k, v in rec.pii_hits.items():
                                pii_totals[k] = pii_totals.get(k, 0) + int(v)

                    if enable_pii_samples and pii_max_samples_per_file > 0:
                        matches = find_pii_matches(sample_text, max_matches=pii_max_samples_per_file)
                        for m in matches:
                            start = int(getattr(m, "start", 0) or 0)
                            end = int(getattr(m, "end", 0) or 0)
                            if end <= start:
                                continue
                            ctx_start = max(0, start - int(pii_context_chars))
                            ctx_end = min(len(sample_text), end + int(pii_context_chars))
                            ctx = sample_text[ctx_start:ctx_end]
                            # Mask other occurrences within context for safer display.
                            ctx = anonymize_pii(ctx, enabled=True, mode="mask").text
                            ctx = redact_secrets(ctx, enabled=True, mode="mask").text
                            if len(ctx) > 2000:
                                ctx = ctx[:2000] + "..."
                            rec.pii_samples.append(
                                {
                                    "kind": str(getattr(m, "kind", "") or "pii"),
                                    "masked": _mask_pii_value(str(getattr(m, "kind", "") or ""), str(getattr(m, "text", "") or "")),
                                    "context": ctx,
                                    "start": start,
                                    "end": end,
                                }
                            )

                if sample_text and enable_secrets:
                    sec = redact_secrets(sample_text, enabled=True, mode="mask")
                    if sec.hits:
                        rec.secrets_hits = {str(k): int(v) for k, v in sec.hits.items() if int(v) > 0}
                        if rec.secrets_hits:
                            rec.findings.append("secrets")
                            finding_counts["secrets"] += 1
                            for k, v in rec.secrets_hits.items():
                                secrets_totals[k] = secrets_totals.get(k, 0) + int(v)

                    if enable_secrets_samples and secrets_max_samples_per_file > 0:
                        matches = find_secret_matches(sample_text, max_matches=secrets_max_samples_per_file)
                        for m in matches:
                            start = int(getattr(m, "start", 0) or 0)
                            end = int(getattr(m, "end", 0) or 0)
                            if end <= start:
                                continue
                            ctx_start = max(0, start - int(secrets_context_chars))
                            ctx_end = min(len(sample_text), end + int(secrets_context_chars))
                            ctx = sample_text[ctx_start:ctx_end]
                            # Mask PII/secrets in context for safer display.
                            ctx = anonymize_pii(ctx, enabled=True, mode="mask").text
                            ctx = redact_secrets(ctx, enabled=True, mode="mask").text
                            if len(ctx) > 2000:
                                ctx = ctx[:2000] + "..."
                            rec.secrets_samples.append(
                                {
                                    "kind": str(getattr(m, "kind", "") or "secret"),
                                    "masked": _mask_secret_value(str(getattr(m, "kind", "") or ""), str(getattr(m, "text", "") or "")),
                                    "context": ctx,
                                    "start": start,
                                    "end": end,
                                }
                            )

                # Optional file hash (expensive; for exact duplicates).
                if compute_file_hash:
                    sha = _safe_hash_file(path, algo="sha256")
                    rec.file_sha256 = sha
                    sha_counts[sha] = sha_counts.get(sha, 0) + 1

            except Exception as exc:  # noqa: BLE001
                errors += 1
                rec.error_message = str(exc)[:200]
                rec.findings.append("parse_failed")
                finding_counts["parse_failed"] += 1

            # Aggregation (best-effort).
            by_type[rec.file_type] = by_type.get(rec.file_type, 0) + 1
            file_sizes.append(int(rec.file_size or 0))
            if int(rec.text_characters or 0) > 0:
                text_lengths.append(int(rec.text_characters))

            # Write JSONL line.
            jf.write(json.dumps(asdict(rec), ensure_ascii=False, separators=(",", ":")))
            jf.write("\n")

            flush_progress(idx)

    # Add exact-dup counts.
    if compute_file_hash and sha_counts:
        dup_total = 0
        for _sha, cnt in sha_counts.items():
            if int(cnt) > 1:
                dup_total += int(cnt)
        if dup_total > 0:
            finding_counts["exact_dup"] = int(dup_total)

    near_dup_path: Path | None = None
    if enable_near_dup and simhash_entries and near_dup_hamming_threshold > 0 and near_dup_max_pairs > 0:
        names = [n for n, _h in simhash_entries]
        hashes = [int(_h) for _n, _h in simhash_entries]
        n_total = len(hashes)

        # LSH banding to avoid O(N^2) comparisons (4x16-bit bands).
        buckets: dict[tuple[int, int], list[int]] = {}
        pairs: list[dict[str, Any]] = []
        parent = list(range(n_total))
        rank = [0] * n_total

        def find(x: int) -> int:
            while parent[x] != x:
                parent[x] = parent[parent[x]]
                x = parent[x]
            return x

        def union(a: int, b: int) -> None:
            ra = find(a)
            rb = find(b)
            if ra == rb:
                return
            if rank[ra] < rank[rb]:
                parent[ra] = rb
            elif rank[ra] > rank[rb]:
                parent[rb] = ra
            else:
                parent[rb] = ra
                rank[ra] += 1

        # Compare candidates as we insert each item.
        for i, h in enumerate(hashes):
            candidates: set[int] = set()
            for band in range(4):
                key = (band, int((h >> (band * 16)) & 0xFFFF))
                existing = buckets.get(key)
                if existing and len(existing) <= 2000:
                    candidates.update(existing)
                buckets.setdefault(key, []).append(i)

            for j in candidates:
                if j == i:
                    continue
                d = _hamming_distance64(h, hashes[j])
                if d <= int(near_dup_hamming_threshold):
                    pairs.append({"a": names[j], "b": names[i], "distance": int(d)})
                    union(i, j)
                    if len(pairs) >= int(near_dup_max_pairs):
                        break
            if len(pairs) >= int(near_dup_max_pairs):
                break

        # Build clusters from union-find (size>=2).
        groups: dict[int, list[int]] = {}
        for idx in range(n_total):
            r = find(idx)
            groups.setdefault(r, []).append(idx)

        clusters: list[dict[str, Any]] = []
        for root_id, members in groups.items():
            if len(members) < 2:
                continue
            clusters.append({"id": str(root_id), "members": [names[i] for i in sorted(members)]})

        clusters.sort(key=lambda c: (-len(c.get("members") or []), str(c.get("id") or "")))
        affected = {m for c in clusters for m in (c.get("members") or [])}
        if affected:
            finding_counts["near_dup"] = int(len(affected))

        near_dup_payload = {
            "threshold": int(near_dup_hamming_threshold),
            "max_pairs": int(near_dup_max_pairs),
            "pairs_returned": int(len(pairs)),
            "clusters_returned": int(len(clusters)),
            "clusters": clusters[:2000],
            "pairs": pairs[:5000],
        }
        near_dup_path = (artifact_root / "near_dups.json").resolve(strict=False)
        near_dup_path.write_text(json.dumps(near_dup_payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

    # Histograms + percentiles.
    file_sizes.sort()
    text_lengths.sort()
    percentiles = {
        "p25": percentile_from_sorted(text_lengths, 25),
        "p50": percentile_from_sorted(text_lengths, 50),
        "p75": percentile_from_sorted(text_lengths, 75),
        "p90": percentile_from_sorted(text_lengths, 90),
        "p99": percentile_from_sorted(text_lengths, 99),
    }

    summary = {
        "dataset_id": str(dataset_id),
        "scan_run_id": str(run.id),
        "generated_at": _now_utc().isoformat(),
        "total_files": int(total),
        "total_size_bytes": int(sum(file_sizes)),
        "by_file_type": {k: int(v) for k, v in sorted(by_type.items(), key=lambda kv: (-kv[1], kv[0]))},
        "file_size_histogram": histogram(file_sizes, FILE_SIZE_BINS),
        "length_percentiles": percentiles,
        "length_histogram": histogram(text_lengths, TEXT_LENGTH_BINS),
        "pdf_scan": {"scanned": int(pdf_scanned), "not_scanned": int(pdf_not_scanned), "unknown": int(pdf_unknown)},
        "pdf_detection": {
            "sample_pages": int(pdf_sample_pages),
            "scan_max_chars_per_page": int(pdf_scan_max_chars),
            "text_min_chars_per_page": int(pdf_text_min_chars),
            "scan_ratio_threshold": float(pdf_scan_ratio_threshold),
        },
        "pii_hits_total": {k: int(v) for k, v in pii_totals.items()},
        "secrets_hits_total": {k: int(v) for k, v in secrets_totals.items()},
        "findings": [
            {
                "key": k,
                "label": v.get("label", k),
                "severity": v.get("severity", "info"),
                "count": int(finding_counts.get(k, 0) or 0),
                "description": v.get("description"),
            }
            for k, v in FINDING_KEY_REASONS.items()
        ],
    }

    run.status = "completed"
    run.progress = 100
    run.finished_at = _now_utc()
    run.summary = summary
    run.artifacts = {
        "files_jsonl": str(jsonl_path),
        "root_path": "[REDACTED]" if redact_paths else str(root),
    }
    if near_dup_path is not None:
        run.artifacts["near_dups_json"] = str(near_dup_path)
    db.commit()

    return {
        "ok": True,
        "files": int(total),
        "errors": int(errors),
    }
