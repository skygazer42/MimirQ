"""
Excel parser (fallback / lightweight).

Used as a best-effort fallback when MarkItDown is unavailable or fails.
Supports .xlsx via openpyxl; .xls uses pandas engines and may fail.

Outputs Markdown tables (best-effort) for better fidelity in downstream preview/RAG.
"""


import io
from pathlib import Path
from typing import Any

from langchain_core.documents import Document

from app.rag.core.logging import get_logger

logger = get_logger(__name__)

_EMPTY_SHEET_MARKDOWN = "_Empty sheet._\n\n"


def _safe_cell(value: object, *, max_chars: int) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\r", " ").replace("\n", " ").strip()
    if max_chars > 0 and len(text) > max_chars:
        return text[: max_chars - 1] + "…"
    return text


def _escape_md_cell(text: str) -> str:
    text = (text or "").replace("\r", " ").replace("\n", " ").strip()
    text = " ".join(text.split())
    # Keep Markdown tables stable.
    return text.replace("|", r"\|")


def _md_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    norm = [r + [""] * (width - len(r)) for r in rows]
    header = norm[0]
    body = norm[1:]
    out: list[str] = []
    out.append("| " + " | ".join(header) + " |")
    out.append("| " + " | ".join(["---"] * width) + " |")
    for r in body:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out).strip()


class ExcelParser:
    """Parse Excel into a row-oriented, chunk-friendly text representation."""

    def __init__(
        self,
        *,
        max_sheets: int = 0,
        max_rows: int = 0,
        max_cols: int = 0,
        max_cell_chars: int = 2000,
    ) -> None:
        self.max_sheets = int(max_sheets or 0)
        self.max_rows = int(max_rows or 0)
        self.max_cols = int(max_cols or 0)
        self.max_cell_chars = int(max_cell_chars or 0)

    def parse(self, file_path: Path) -> list[Document]:
        ext = file_path.suffix.lower()

        if ext == ".xlsx":
            return self._parse_xlsx(file_path)

        # Best-effort for .xls using pandas; may require extra engines.
        return self._parse_via_pandas(file_path)

    def _sheet_limits(self, ws: Any) -> tuple[int, int]:
        max_rows = self.max_rows if self.max_rows > 0 else int(getattr(ws, "max_row", 0) or 0)
        max_cols = self.max_cols if self.max_cols > 0 else int(getattr(ws, "max_column", 0) or 0)
        return max(0, int(max_rows)), max(0, int(max_cols))

    @staticmethod
    def _build_cell_grid(ws: Any, *, max_rows: int, max_cols: int) -> list[list[object]]:
        grid: list[list[object]] = [[None for _ in range(max_cols)] for _ in range(max_rows)]
        for r in range(1, max_rows + 1):
            for c in range(1, max_cols + 1):
                try:
                    grid[r - 1][c - 1] = ws.cell(row=r, column=c).value
                except Exception:
                    grid[r - 1][c - 1] = None
        return grid

    @staticmethod
    def _merged_ranges(ws: Any) -> list[object]:
        try:
            return list(getattr(getattr(ws, "merged_cells", None), "ranges", []) or [])
        except Exception:
            return []

    @staticmethod
    def _fill_merged_ranges(ws: Any, *, grid: list[list[object]], max_rows: int, max_cols: int) -> None:
        for merged in ExcelParser._merged_ranges(ws):
            try:
                bounds = getattr(merged, "bounds", None)
                if not isinstance(bounds, tuple) or len(bounds) != 4:
                    raise ValueError("missing_bounds")
                min_col, min_row, max_col, max_row = bounds
            except Exception:
                get_logger(__name__).debug("Skipping item after non-critical exception", exc_info=True)
                continue
            if min_row > max_rows or min_col > max_cols:
                continue
            try:
                top_left = ws.cell(row=min_row, column=min_col).value
            except Exception:
                top_left = None
            for r in range(min_row, min(max_row, max_rows) + 1):
                for c in range(min_col, min(max_col, max_cols) + 1):
                    grid[r - 1][c - 1] = top_left

    @staticmethod
    def _row_has_data(row: list[object]) -> bool:
        return any(str(v).strip() for v in row if v is not None)

    @classmethod
    def _trim_grid(cls, grid: list[list[object]], *, max_cols: int) -> list[list[object]]:
        last_row = 0
        for index, row in enumerate(grid, start=1):
            if cls._row_has_data(row):
                last_row = index
        if last_row <= 0:
            return []
        trimmed = grid[:last_row]

        last_col = 0
        for col_index in range(1, max_cols + 1):
            if any(row[col_index - 1] is not None and str(row[col_index - 1]).strip() for row in trimmed):
                last_col = col_index
        if last_col <= 0:
            return []
        return [row[:last_col] for row in trimmed]

    def _markdown_rows(self, grid: list[list[object]]) -> tuple[list[list[str]], int]:
        rows_md: list[list[str]] = []
        emitted_rows = 0
        for row in grid:
            cells = [_escape_md_cell(_safe_cell(value, max_chars=self.max_cell_chars)) for value in (row or [])]
            if not any(cells):
                continue
            emitted_rows += 1
            rows_md.append(cells)
        return rows_md, emitted_rows

    def _sheet_markdown(self, ws: Any) -> tuple[str, int, bool]:
        max_rows, max_cols = self._sheet_limits(ws)
        if max_rows <= 0 or max_cols <= 0:
            return _EMPTY_SHEET_MARKDOWN, 0, False

        grid = self._build_cell_grid(ws, max_rows=max_rows, max_cols=max_cols)
        self._fill_merged_ranges(ws, grid=grid, max_rows=max_rows, max_cols=max_cols)
        grid = self._trim_grid(grid, max_cols=max_cols)
        if not grid:
            return _EMPTY_SHEET_MARKDOWN, 0, False

        rows_md, emitted_rows = self._markdown_rows(grid)
        if not rows_md:
            return _EMPTY_SHEET_MARKDOWN, 0, False

        try:
            truncated = (
                (self.max_rows > 0 and int(getattr(ws, "max_row", 0) or 0) > self.max_rows)
                or (self.max_cols > 0 and int(getattr(ws, "max_column", 0) or 0) > self.max_cols)
            )
        except Exception as exc:
            logger.debug("Ignoring Excel truncation metadata failure: %s", exc)
            truncated = False
        return _md_table(rows_md) + "\n\n", emitted_rows, truncated

    def _parse_xlsx(self, file_path: Path) -> list[Document]:
        from openpyxl import load_workbook  # type: ignore

        # read_only=False so we can inspect merged cells reliably.
        wb = load_workbook(filename=str(file_path), read_only=False, data_only=True)
        try:
            sheet_names = list(getattr(wb, "sheetnames", []) or [])
            limited = sheet_names[: self.max_sheets] if self.max_sheets > 0 else sheet_names

            out = io.StringIO()
            out.write(f"Excel: {file_path.name}\n")
            out.write(f"Sheets: {', '.join(limited) if limited else '(none)'}\n\n")

            emitted_rows = 0
            truncated_any = False
            for sheet_name in limited:
                ws = wb[sheet_name]
                out.write(f"## Sheet: {sheet_name}\n\n")
                sheet_markdown, sheet_rows, sheet_truncated = self._sheet_markdown(ws)
                out.write(sheet_markdown)
                emitted_rows += sheet_rows
                truncated_any = truncated_any or sheet_truncated

            metadata = {
                "source": str(file_path.name),
                "file_type": "xlsx",
                "excel_sheets": limited,
                "excel_rows_emitted": int(emitted_rows),
            }
            if truncated_any:
                metadata["excel_truncated"] = True
            return [Document(page_content=out.getvalue(), metadata=metadata)]
        finally:
            try:
                wb.close()
            except Exception as exc:
                logger.debug("Ignoring Excel workbook close failure: %s", exc)

    def _parse_via_pandas(self, file_path: Path) -> list[Document]:
        import pandas as pd  # type: ignore

        # Read a small preview; pandas will select an engine based on extension.
        try:
            df = pd.read_excel(str(file_path), nrows=self.max_rows if self.max_rows > 0 else None)
        except Exception as exc:
            raise RuntimeError(f"Excel preview failed: {exc}") from exc

        # Limit columns.
        if self.max_cols > 0 and getattr(df, "shape", (0, 0))[1] > self.max_cols:
            df = df.iloc[:, : self.max_cols]

        # Render as a Markdown table (best-effort), similar to the .xlsx path.
        cols: list[str] = []
        try:
            cols = [str(c) for c in df.columns]
        except Exception:
            cols = []

        rows_md: list[list[str]] = []
        if cols:
            rows_md.append([_escape_md_cell(_safe_cell(c, max_chars=self.max_cell_chars)) for c in cols])
        for row in df.itertuples(index=False, name=None):
            values = [_escape_md_cell(_safe_cell(v, max_chars=self.max_cell_chars)) for v in (row or ())]
            if any(values):
                rows_md.append(values)

        out = io.StringIO()
        out.write(f"Excel: {file_path.name}\n\n")
        if rows_md:
            out.write(_md_table(rows_md))
            out.write("\n")
        else:
            out.write("_Empty sheet._\n")

        cols: list[str] | None = None
        try:
            cols = [str(c) for c in df.columns]
        except Exception:
            cols = None

        metadata = {
            "source": str(file_path.name),
            "file_type": file_path.suffix.lstrip("."),
            "excel_rows_emitted": int(getattr(df, "shape", (0, 0))[0]),
        }
        if cols:
            metadata["excel_columns"] = cols[:200]
        return [Document(page_content=out.getvalue(), metadata=metadata)]
