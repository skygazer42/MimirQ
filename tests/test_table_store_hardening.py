from __future__ import annotations

import sqlite3
import uuid
from pathlib import Path

import pytest


def _write_demo_xlsx(path: Path, *, sheets: int) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    # Workbook starts with 1 sheet.
    for i in range(max(0, int(sheets) - 1)):
        wb.create_sheet(title=f"s{i+1}")
    for idx, name in enumerate(list(getattr(wb, "sheetnames", []) or [])):
        ws = wb[name]
        ws["A1"] = "a"
        ws["B1"] = "b"
        ws["A2"] = idx
        ws["B2"] = idx + 1
    wb.save(path)


def test_table_store_import_caps_sheet_count(monkeypatch, tmp_path: Path) -> None:
    from app.core.config import settings
    from app.services.table_store import table_store_path
    from app.services.table_store_service import import_table_document

    tenant_id = uuid.uuid4()
    dataset_id = uuid.uuid4()
    document_id = uuid.uuid4()

    store_dir = tmp_path / "table_store"
    monkeypatch.setattr(settings, "TABLE_STORE_DIR", str(store_dir), raising=False)
    monkeypatch.setattr(settings, "TABLE_STORE_MAX_SHEETS", 2, raising=False)

    xlsx = tmp_path / "demo.xlsx"
    _write_demo_xlsx(xlsx, sheets=3)

    assets = import_table_document(
        tenant_id=tenant_id,
        dataset_id=dataset_id,
        document_id=document_id,
        file_path=xlsx,
        max_rows=1000,
        max_cols=1000,
        sample_rows=0,
    )
    assert len(assets) == 2
    assert all(a.truncated for a in assets)

    db_path = table_store_path(tenant_id=tenant_id, dataset_id=dataset_id, document_id=document_id)
    assert db_path.exists()
    conn = sqlite3.connect(str(db_path))
    try:
        names = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    finally:
        conn.close()
    assert "sheet_0" in names
    assert "sheet_1" in names
    assert "sheet_2" not in names


def test_table_store_reimport_drops_stale_tables_when_unlink_fails(monkeypatch, tmp_path: Path) -> None:
    from app.core.config import settings
    from app.services.table_store import table_store_path
    from app.services.table_store_service import import_table_document

    tenant_id = uuid.uuid4()
    dataset_id = uuid.uuid4()
    document_id = uuid.uuid4()

    store_dir = tmp_path / "table_store"
    monkeypatch.setattr(settings, "TABLE_STORE_DIR", str(store_dir), raising=False)
    monkeypatch.setattr(settings, "TABLE_STORE_MAX_SHEETS", 0, raising=False)

    xlsx1 = tmp_path / "two_sheets.xlsx"
    _write_demo_xlsx(xlsx1, sheets=2)
    import_table_document(
        tenant_id=tenant_id,
        dataset_id=dataset_id,
        document_id=document_id,
        file_path=xlsx1,
        max_rows=1000,
        max_cols=1000,
        sample_rows=0,
    )

    db_path = table_store_path(tenant_id=tenant_id, dataset_id=dataset_id, document_id=document_id)
    assert db_path.exists()

    # Force Path.unlink to fail for this specific sqlite file so import has to do the "drop sheet_* tables" fallback.
    import pathlib

    orig_unlink = pathlib.Path.unlink

    def _fake_unlink(self: pathlib.Path, *args, **kwargs):  # noqa: ANN001
        if self.resolve(strict=False) == db_path.resolve(strict=False):
            raise PermissionError("simulated unlink failure")
        return orig_unlink(self, *args, **kwargs)

    monkeypatch.setattr(pathlib.Path, "unlink", _fake_unlink, raising=True)

    xlsx2 = tmp_path / "one_sheet.xlsx"
    _write_demo_xlsx(xlsx2, sheets=1)
    assets = import_table_document(
        tenant_id=tenant_id,
        dataset_id=dataset_id,
        document_id=document_id,
        file_path=xlsx2,
        max_rows=1000,
        max_cols=1000,
        sample_rows=0,
    )
    assert assets and {a.sheet_index for a in assets} == {0}

    conn = sqlite3.connect(str(db_path))
    try:
        names = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
    finally:
        conn.close()
    assert "sheet_0" in names
    assert "sheet_1" not in names


def test_table_query_rejects_schema_table_references(monkeypatch, tmp_path: Path) -> None:
    from app.core.config import settings
    from app.services.table_store_service import import_table_document, run_table_query

    tenant_id = uuid.uuid4()
    dataset_id = uuid.uuid4()
    document_id = uuid.uuid4()

    store_dir = tmp_path / "table_store"
    monkeypatch.setattr(settings, "TABLE_STORE_DIR", str(store_dir), raising=False)

    xlsx = tmp_path / "demo.xlsx"
    _write_demo_xlsx(xlsx, sheets=1)
    import_table_document(
        tenant_id=tenant_id,
        dataset_id=dataset_id,
        document_id=document_id,
        file_path=xlsx,
        max_rows=1000,
        max_cols=1000,
        sample_rows=0,
    )

    table_id = f"doc:{document_id}:sheet:0"
    with pytest.raises(ValueError) as exc:
        run_table_query(
            tenant_id=tenant_id,
            dataset_id=dataset_id,
            table_id=table_id,
            sql="SELECT name FROM sqlite_master LIMIT 5",
            max_rows=10,
            max_cols=10,
            max_bytes=10000,
        )
    assert "schema_table_reference_not_allowed" in str(exc.value)

