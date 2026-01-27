from __future__ import annotations

from pathlib import Path

from app.services.table_routing import decide_table_route


def test_table_route_auto_route_disabled_defaults_to_tag(tmp_path: Path):
    p = tmp_path / "a.csv"
    p.write_text("a,b\n1,2\n", encoding="utf-8")
    d = decide_table_route(
        p,
        auto_route=False,
        file_bytes_threshold=0,
        row_threshold=5000,
        col_threshold=80,
        sheet_threshold=5,
    )
    assert d.route == "tag"
    assert d.reason == "auto_route_disabled"


def test_table_route_csv_small_goes_rag(tmp_path: Path):
    p = tmp_path / "small.csv"
    p.write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
    d = decide_table_route(
        p,
        auto_route=True,
        file_bytes_threshold=0,
        row_threshold=100,
        col_threshold=80,
        sheet_threshold=0,
    )
    assert d.route == "rag"


def test_table_route_csv_rows_threshold_goes_tag(tmp_path: Path):
    p = tmp_path / "rows.csv"
    lines = ["a,b"] + ["1,2"] * 6000
    p.write_text("\n".join(lines) + "\n", encoding="utf-8")
    d = decide_table_route(
        p,
        auto_route=True,
        file_bytes_threshold=0,
        row_threshold=5000,
        col_threshold=0,
        sheet_threshold=0,
    )
    assert d.route == "tag"
    assert d.reason == "rows_threshold"


def test_table_route_csv_cols_threshold_goes_tag(tmp_path: Path):
    p = tmp_path / "cols.csv"
    header = ",".join([f"c{i}" for i in range(100)])
    row = ",".join(["1"] * 100)
    p.write_text(header + "\n" + row + "\n", encoding="utf-8")
    d = decide_table_route(
        p,
        auto_route=True,
        file_bytes_threshold=0,
        row_threshold=10_000,
        col_threshold=80,
        sheet_threshold=0,
    )
    assert d.route == "tag"
    assert d.reason == "cols_threshold"


def test_table_route_file_bytes_threshold_goes_tag(tmp_path: Path):
    p = tmp_path / "big.csv"
    # Keep content simple; we only need file size to exceed threshold.
    p.write_text("x" * 2048, encoding="utf-8")
    d = decide_table_route(
        p,
        auto_route=True,
        file_bytes_threshold=1024,
        row_threshold=0,
        col_threshold=0,
        sheet_threshold=0,
    )
    assert d.route == "tag"
    assert d.reason == "file_bytes_threshold"


def test_table_route_xlsx_small_goes_rag(tmp_path: Path):
    from openpyxl import Workbook

    p = tmp_path / "small.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "a"
    ws["B1"] = "b"
    ws["A2"] = 1
    ws["B2"] = 2
    wb.save(p)

    d = decide_table_route(
        p,
        auto_route=True,
        file_bytes_threshold=0,
        row_threshold=5000,
        col_threshold=80,
        sheet_threshold=5,
    )
    assert d.route == "rag"


def test_table_route_xlsx_rows_threshold_goes_tag(tmp_path: Path):
    from openpyxl import Workbook

    p = tmp_path / "rows.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.cell(row=6000, column=1, value=1)  # sets max_row=6000 without filling all cells
    wb.save(p)

    d = decide_table_route(
        p,
        auto_route=True,
        file_bytes_threshold=0,
        row_threshold=5000,
        col_threshold=0,
        sheet_threshold=0,
    )
    assert d.route == "tag"
    assert d.reason == "rows_threshold"


def test_table_route_xlsx_cols_threshold_goes_tag(tmp_path: Path):
    from openpyxl import Workbook

    p = tmp_path / "cols.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=100, value="x")  # sets max_column=100
    wb.save(p)

    d = decide_table_route(
        p,
        auto_route=True,
        file_bytes_threshold=0,
        row_threshold=0,
        col_threshold=80,
        sheet_threshold=0,
    )
    assert d.route == "tag"
    assert d.reason == "cols_threshold"


def test_table_route_xlsx_sheet_threshold_goes_tag(tmp_path: Path):
    from openpyxl import Workbook

    p = tmp_path / "sheets.xlsx"
    wb = Workbook()
    # Workbook starts with 1 sheet; add 5 more -> 6 total
    for i in range(5):
        wb.create_sheet(title=f"s{i}")
    wb.save(p)

    d = decide_table_route(
        p,
        auto_route=True,
        file_bytes_threshold=0,
        row_threshold=0,
        col_threshold=0,
        sheet_threshold=5,
    )
    assert d.route == "tag"
    assert d.reason == "sheet_threshold"


def test_table_route_xlsx_corrupt_falls_back_to_rag(tmp_path: Path):
    p = tmp_path / "bad.xlsx"
    p.write_bytes(b"not a real xlsx")
    d = decide_table_route(
        p,
        auto_route=True,
        file_bytes_threshold=0,
        row_threshold=5000,
        col_threshold=80,
        sheet_threshold=5,
    )
    assert d.route == "rag"
    assert d.reason in {"shape_unknown", "below_threshold"}


def test_table_route_xls_legacy_defaults_to_rag(tmp_path: Path):
    p = tmp_path / "legacy.xls"
    p.write_bytes(b"fake xls")
    d = decide_table_route(
        p,
        auto_route=True,
        file_bytes_threshold=0,
        row_threshold=5000,
        col_threshold=80,
        sheet_threshold=5,
    )
    assert d.route == "rag"
    assert d.reason == "legacy_xls_no_signal"

