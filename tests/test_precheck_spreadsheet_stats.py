from __future__ import annotations

from pathlib import Path


def test_precheck_xlsx_spreadsheet_stats_includes_cols(tmp_path: Path) -> None:
    from openpyxl import Workbook

    from app.services.dataset_precheck_scan_runner import _xlsx_spreadsheet_stats

    p = tmp_path / "demo.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "a"
    ws["B1"] = "b"
    ws["C1"] = "c"
    ws.cell(row=10, column=3, value=1)  # sets max_row=10/max_column=3
    wb.save(p)

    stats = _xlsx_spreadsheet_stats(p)
    assert isinstance(stats, dict)
    assert stats["row_count"] == 10
    assert stats["col_count"] == 3
    assert stats["sheet_count"] == 1
    assert stats["estimated_rows"] is False
    assert stats["estimated_cols"] is False

